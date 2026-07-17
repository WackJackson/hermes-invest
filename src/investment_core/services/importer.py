from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from investment_core.models import HoldingRecord


REQUIRED_COLUMNS = [
    "symbol",
    "market",
    "name",
    "quantity",
    "avg_cost",
    "currency",
    "account",
    "open_date",
]


def _normalize_record(row: dict[str, object]) -> HoldingRecord:
    normalized = {key: row[key] for key in REQUIRED_COLUMNS}
    normalized["market"] = str(normalized["market"]).upper()
    normalized["symbol"] = str(normalized["symbol"]).strip().upper()
    normalized["currency"] = str(normalized["currency"]).strip().upper()
    normalized["name"] = str(normalized["name"]).strip()
    normalized["account"] = str(normalized["account"]).strip()
    normalized["open_date"] = str(normalized["open_date"]).strip()
    normalized["quantity"] = float(normalized["quantity"])
    normalized["avg_cost"] = float(normalized["avg_cost"])
    return HoldingRecord(**normalized)


def _ensure_columns(columns: list[str]) -> None:
    missing = [column for column in REQUIRED_COLUMNS if column not in columns]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")


def _parse_csv(content: bytes) -> list[HoldingRecord]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV file is empty")
    _ensure_columns(reader.fieldnames)
    return [_normalize_record(row) for row in reader]


def _parse_excel(content: bytes) -> list[HoldingRecord]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")
    headers = [str(cell).strip() for cell in rows[0]]
    _ensure_columns(headers)
    records = []
    for values in rows[1:]:
        if values is None or all(cell is None for cell in values):
            continue
        row = {headers[index]: values[index] for index in range(len(headers))}
        records.append(_normalize_record(row))
    return records


def parse_holdings_upload(filename: str, content: bytes) -> list[HoldingRecord]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(content)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_excel(content)
    raise ValueError("Only CSV and XLSX files are supported in MVP")
